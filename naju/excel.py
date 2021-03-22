from datetime import datetime

import xlsxwriter as wr
import os

from typing import Optional

from flask import current_app
from naju.database import get_db
from xlsxwriter.worksheet import (Worksheet, cell_number_tuple, cell_string_tuple)
from openpyxl import load_workbook


def get_column_width(worksheet: Worksheet, column: int) -> Optional[int]:
    """Get the max column width in a `Worksheet` column."""
    strings = getattr(worksheet, '_ts_all_strings', None)
    if strings is None:
        strings = worksheet._ts_all_strings = sorted(
            worksheet.str_table.string_table,
            key=worksheet.str_table.string_table.__getitem__)
    lengths = set()
    for row_id, colums_dict in worksheet.table.items():  # type: int, dict
        data = colums_dict.get(column)
        if not data:
            continue
        if type(data) is cell_string_tuple:
            iter_length = len(strings[data.string])
            if not iter_length:
                continue
            lengths.add(iter_length)
            continue
        if type(data) is cell_number_tuple:
            iter_length = len(str(data.number))
            if not iter_length:
                continue
            lengths.add(iter_length)
    if not lengths:
        return None
    return max(lengths)


def set_column_autowidth(worksheet: Worksheet, column: int):
    """
    Set the width automatically on a column in the `Worksheet`.
    !!! Make sure you run this function AFTER having all cells filled in
    the worksheet!
    """
    maxwidth = get_column_width(worksheet=worksheet, column=column)
    if maxwidth is None:
        return
    worksheet.set_column(first_col=column, last_col=column, width=maxwidth + 10)


def load_table():
    print('Hi')
    path = os.path.join(current_app.instance_path, 'assets', 'uploads', 'workbook.xlsx')
    print(path)
    wb = load_workbook(path)

    db = get_db()

    art = db.execute('SELECT * FROM tree_param_type WHERE name = ?', ("Art", )).fetchone()
    sorte = db.execute('SELECT * FROM tree_param_type WHERE name = ?', ("Sorte",)).fetchone()
    pflanz = db.execute('SELECT * FROM tree_param_type WHERE name = ?', ("Pflanzung",)).fetchone()
    u = db.execute('SELECT * FROM tree_param_type WHERE name = ?', ("Umfang in cm",)).fetchone()
    h = db.execute('SELECT * FROM tree_param_type WHERE name = ?', ("Höhe in cm",)).fetchone()
    vit = db.execute('SELECT * FROM tree_param_type WHERE name = ?', ("Vital",)).fetchone()

    params = [art, sorte, pflanz, u, h, vit]

    for sheet in wb.worksheets:
        area = db.execute('SELECT * FROM area WHERE name = ?', (sheet.title, )).fetchone()
        if area is not None:
            for row in sheet.rows:
                digit = str(row[0].value).replace('.', '')
                if str(digit).isnumeric():
                    if db.execute('SELECT * FROM tree t, area a '
                                  'WHERE t.area_id = a.id AND t.area_id = ? AND t.number = ?',
                                  (area['id'], int(digit), )).fetchone() is None:

                        db.execute("INSERT INTO tree (number, area_id) VALUES (?, ?)",
                                   (int(digit), area['id'], ))
                        print(row[0].value)
                    tree = db.execute('SELECT * FROM tree WHERE number = ? AND area_id = ?', (int(digit), area['id'], ))

                    row_num = 1

                    for param in params:
                        if row[row_num].value is not None and param is not None:
                            if db.execute('SELECT * FROM tree_param '
                                          'WHERE tree_id = ? AND param_id = ?',
                                          (tree['id'], param['id'])).fetchone() is None:
                                db.execute('INSERT INTO tree_param (tree_id, param_id, value) VALUES (?, ?, ?)',
                                           (tree['id'], param['id'], row[row_num].value))
                        row_num += 1

    wb.close()
    db.commit()


def create_table(type, filter):
    path = os.path.join(current_app.instance_path, 'tables')
    file = os.path.join(path, "type-filter.xlsx")

    os.makedirs(path, exist_ok=True)

    workbook = wr.Workbook(file)

    workbook.set_properties({
        'title': 'Baumbestand NAJU Essen/Mülheim (' + filter + ')',
        'author': 'Fabius Mettner',
        'company': 'NAJU - Essen/Mülheim',
        'category': 'Baumbestand',
        'keywords': 'NAJU, Baumbestand',
        'created': datetime.now(),
        'comments': 'Created with Python and XlsxWriter'})

    merge_format = workbook.add_format({
        'bold': True,
        'border': 5,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '#D7E4BC',
        'font_size': 24,
        'bg_color': 'gray'
    })

    header_format = workbook.add_format({
        'bold': True,
        'border': 2,
        'align': 'center',
        'font_size': 16,
        'bg_color': 'gray'
    })

    format_1 = workbook.add_format({
        'bold': False,
        'border': 1,
        'align': 'left',
        'font_size': 12,
        'bg_color': 'white'
    })

    format_2 = workbook.add_format({
        'bold': False,
        'border': 1,
        'align': 'left',
        'font_size': 12,
        'bg_color': '#9b9b9b'
    })

    db = get_db()

    areas = db.execute('SELECT * FROM area').fetchall()

    for area in areas:
        try:
            if type == "Fläche" and str(area['name']).upper().index(str(filter).upper()) < 0:
                continue
        except ValueError:
            continue
        worksheet = workbook.add_worksheet(area['name'])

        row = 0
        col = 0

        row += 2

        worksheet.write(row, col, 'Nummer', header_format)

        col += 1

        params = db.execute('SELECT * FROM tree_param_type ORDER BY name')

        for param in params:
            worksheet.write(row, col, param['name'], header_format)
            col += 1

        row += 1

        worksheet.merge_range(0, 0, 0, col - 1, area['name'], merge_format)

        trees = db.execute('SELECT * FROM tree WHERE area_id=? ORDER BY number', (area['id'],)).fetchall()

        for tree in trees:
            jump = False
            params = db.execute('SELECT * FROM tree_param tp, tree_param_type tpt '
                                'WHERE tp.param_id = tpt.id AND tree_id=? ORDER BY name', (tree['id'],)).fetchall()

            for param in params:
                try:
                    if param['name'] == type and \
                            str(param['value']).upper().index(str(filter).upper()) < 0:
                        jump = True
                except ValueError:
                    jump = True
            try:
                if jump or (type == 'Nummer' and str(tree['number']).upper().index(str(filter).upper()) < 0):
                    continue
            except ValueError:
                continue
            col = 0
            if row % 2 == 1:
                worksheet.write(row, col, tree['number'], format_1)
            else:
                worksheet.write(row, col, tree['number'], format_2)
            col += 1

            for param in params:
                if row % 2 == 1:
                    worksheet.write(row, col, param['value'], format_1)
                else:
                    worksheet.write(row, col, param['value'], format_2)
                col += 1
            row += 1

        for i in range(0, col):
            set_column_autowidth(worksheet, i)

    workbook.close()

    return file
